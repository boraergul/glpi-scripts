#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GLPI SLA Compliance Report
==========================
Entity bazlı SLA uyum raporu oluşturur.
Gecikmiş SLA'lara sahip ticketları entity'lere göre kırarak raporlar.

Kullanım:
    python sla_compliance_report.py --start-date 2024-01-01 --end-date 2024-12-31
    python sla_compliance_report.py --start-date 2024-01-01 --end-date 2024-12-31 --export csv
    python sla_compliance_report.py --start-date 2024-01-01 --end-date 2024-12-31 --export excel
"""

import requests
import urllib3
import json
import sys
import os
import argparse
from datetime import datetime
from collections import defaultdict

# SSL uyarılarını kapat
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Config dosyası yolları
CONFIG_PATHS = [
    'config.json',
    '../config/config.json',
    '../../config/config.json'
]

def load_config():
    """Config dosyasını yükle"""
    for config_path in CONFIG_PATHS:
        full_path = os.path.join(os.path.dirname(__file__), config_path)
        if os.path.exists(full_path):
            with open(full_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    raise FileNotFoundError("config.json bulunamadı!")

def init_session(config):
    """GLPI API session başlat"""
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f"user_token {config['GLPI_USER_TOKEN']}",
        'App-Token': config['GLPI_APP_TOKEN']
    }
    
    response = requests.get(
        f"{config['GLPI_URL']}/initSession",
        headers=headers,
        verify=False,
        timeout=30
    )
    
    if response.status_code == 200:
        session_token = response.json()['session_token']
        print(f"✓ GLPI session başlatıldı")
        return session_token
    else:
        raise Exception(f"Session başlatılamadı: {response.text}")

def kill_session(config, session_token):
    """GLPI API session kapat"""
    headers = {
        'Content-Type': 'application/json',
        'Session-Token': session_token,
        'App-Token': config['GLPI_APP_TOKEN']
    }
    
    requests.get(
        f"{config['GLPI_URL']}/killSession",
        headers=headers,
        verify=False,
        timeout=30
    )
    print("✓ Session kapatıldı")

def fetch_all_entities(config, session_token):
    """Tüm entity'leri çek"""
    headers = {
        'Content-Type': 'application/json',
        'Session-Token': session_token,
        'App-Token': config['GLPI_APP_TOKEN']
    }
    
    entities = {}
    range_start = 0
    range_size = 1000
    
    print("Entity'ler çekiliyor...")
    
    while True:
        response = requests.get(
            f"{config['GLPI_URL']}/Entity",
            headers=headers,
            params={'range': f'{range_start}-{range_start + range_size - 1}'},
            verify=False,
            timeout=30
        )
        
        if response.status_code == 200:
            batch = response.json()
            if not batch:
                break
            
            for entity in batch:
                # Entity ismini temizle
                clean_name = entity['completename']
                if clean_name.startswith('Root entity > '):
                    clean_name = clean_name.replace('Root entity > ', '', 1)
                elif clean_name.startswith('Root Entity > '):
                    clean_name = clean_name.replace('Root Entity > ', '', 1)
                elif clean_name.startswith('Ultron Bilişim > '):
                    clean_name = clean_name.replace('Ultron Bilişim > ', '', 1)
                entities[entity['id']] = clean_name
            
            range_start += range_size
        else:
            break
    
    print(f"✓ {len(entities)} entity bulundu")
    return entities

def fetch_all_slas(config, session_token):
    """Tüm SLA tanımlarını çek"""
    headers = {
        'Content-Type': 'application/json',
        'Session-Token': session_token,
        'App-Token': config['GLPI_APP_TOKEN']
    }
    
    slas = {}
    range_start = 0
    range_size = 1000
    
    print("SLA tanımları çekiliyor...")
    
    while True:
        response = requests.get(
            f"{config['GLPI_URL']}/SLA",
            headers=headers,
            params={'range': f'{range_start}-{range_start + range_size - 1}'},
            verify=False,
            timeout=30
        )
        
        if response.status_code == 200:
            batch = response.json()
            if not batch:
                break
            
            for sla in batch:
                slas[sla['id']] = {
                    'name': sla['name'],
                    'type': sla['type'],  # 0=TTR, 1=TTO
                    'number_time': sla.get('number_time', 0),
                    'definition_time': sla.get('definition_time', 'hour')  # day, hour, minute
                }
            
            range_start += range_size
        else:
            break
    
    print(f"✓ {len(slas)} SLA tanımı bulundu")
    return slas

def fetch_tickets(config, session_token, start_date, end_date):
    """Tüm ticketları çek"""
    headers = {
        'Content-Type': 'application/json',
        'Session-Token': session_token,
        'App-Token': config['GLPI_APP_TOKEN']
    }
    
    tickets = []
    range_start = 0
    range_size = 1000
    
    print(f"Ticketlar çekiliyor ({start_date} - {end_date})...")
    
    # Tüm ticketları çek ve filtrele
    while True:
        response = requests.get(
            f"{config['GLPI_URL']}/Ticket",
            headers=headers,
            params={
                'range': f'{range_start}-{range_start + range_size - 1}',
                'expand_dropdowns': 'false',
                'with_devices': 'false',
                'with_disks': 'false',
                'with_softwares': 'false',
                'with_connections': 'false',
                'with_networkports': 'false',
                'with_infocoms': 'false',
                'with_contracts': 'false',
                'with_documents': 'false',
                'with_tickets': 'false',
                'with_problems': 'false',
                'with_changes': 'false',
                'with_notes': 'false',
                'with_logs': 'false'
            },
            verify=False,
            timeout=30
        )
        
        if response.status_code == 200:
            try:
                batch = response.json()
                if not batch:
                    break
                
                # Tarih filtresi uygula
                for ticket in batch:
                    ticket_date = ticket.get('date', '')[:10]  # YYYY-MM-DD formatına çevir
                    
                    # Tarih kontrolü
                    if ticket_date < start_date or ticket_date > end_date:
                        continue
                    
                    tickets.append(ticket)
                
                range_start += range_size
                
                if len(batch) < range_size:
                    break
                    
            except Exception as e:
                print(f"⚠ JSON parse hatası: {e}")
                print(f"Response: {response.text[:200]}")
                break
        elif response.status_code == 206:
            # Partial content, devam et
            try:
                batch = response.json()
                if not batch:
                    break
                
                for ticket in batch:
                    ticket_date = ticket.get('date', '')[:10]
                    if ticket_date >= start_date and ticket_date <= end_date:
                        tickets.append(ticket)
                
                range_start += range_size
            except:
                break
        else:
            print(f"⚠ API Hatası: {response.status_code}")
            break
    
    print(f"✓ {len(tickets)} adet ticket bulundu")
    return tickets

def analyze_sla_compliance(tickets, entities, slas):
    """SLA uyumunu analiz et"""
    print("\nSLA uyumu analiz ediliyor...")
    
    # Entity bazlı istatistikler
    stats = defaultdict(lambda: {
        'total': 0,      # Toplam Ticket
        'sla_total': 0,  # SLA'lı Ticket
        'sla_ok': 0,
        'sla_violated': 0,
        'sla_active': 0,
        'tickets': []
    })
    
    for ticket in tickets:
        entity_id = ticket.get('entities_id', 0)
        entity_name = entities.get(entity_id, f"Unknown ({entity_id})")
        
        stats[entity_name]['total'] += 1
        
        # SLA bilgilerini al
        sla_ttr_id = ticket.get('slas_id_ttr', 0)
        sla_tto_id = ticket.get('slas_id_tto', 0)
        
        # SLA var mı kontrol et
        has_sla = False
        if (sla_ttr_id and sla_ttr_id in slas) or (sla_tto_id and sla_tto_id in slas):
            has_sla = True
            stats[entity_name]['sla_total'] += 1
        
        # Eğer SLA yoksa diğer kontrollere gerek yok, ama ticket listesine ekleyelim mi?
        # Kullanıcı talebine göre detayda sadece SLA'lılar mı isteniyor?
        # Şimdilik sadece SLA'lıları detay listesine ekleyip, sayacı artıralım.
        # Ama döngü SLA varsa devam etmeli.
        
        if not has_sla:
            continue
            
        # Buradan sonrası sadece SLA'lı ticketlar için çalışır
        
        # Gerçek çözüm süresi (saniye cinsinden)
        # GLPI API bazen string olarak döndürüyor, int'e çeviriyoruz
        try:
            solve_delay_stat = int(ticket.get('solve_delay_stat') or 0)
        except (ValueError, TypeError):
            solve_delay_stat = 0
        
        try:
            time_to_own = int(ticket.get('time_to_own') or 0)
        except (ValueError, TypeError):
            time_to_own = 0
        
        # SLA ihlali kontrolü
        has_violation = False
        violation_details = []
        
        # TTR (Time To Resolve) kontrolü
        if sla_ttr_id and sla_ttr_id in slas:
            sla_ttr = slas[sla_ttr_id]
            sla_limit_seconds = convert_to_seconds(
                sla_ttr['number_time'], 
                sla_ttr['definition_time']
            )
            
            # 1. Çözülmüş ticketlarda gecikme kontrolü
            if solve_delay_stat > 0 and solve_delay_stat > sla_limit_seconds:
                has_violation = True
                delay = solve_delay_stat - sla_limit_seconds
                violation_details.append({
                    'type': 'TTR',
                    'sla_name': sla_ttr['name'],
                    'limit': sla_limit_seconds,
                    'actual': solve_delay_stat,
                    'delay': delay,
                    'status': 'solved'
                })
            
            # 2. Henüz çözülmemiş (Aktif) ticketlarda gecikme kontrolü
            # Status < 5 (Solved) ve solve_delay_stat yoksa
            elif ticket.get('status', 0) < 5 and solve_delay_stat == 0:
                time_to_resolve_str = ticket.get('time_to_resolve')
                if time_to_resolve_str:
                    try:
                        ttr_date = datetime.strptime(time_to_resolve_str, '%Y-%m-%d %H:%M:%S')
                        current_time = datetime.now()
                        
                        if current_time > ttr_date:
                            has_violation = True
                            delay_seconds = (current_time - ttr_date).total_seconds()
                            violation_details.append({
                                'type': 'TTR',
                                'sla_name': sla_ttr['name'],
                                'limit': sla_limit_seconds,
                                'actual': int(delay_seconds), # Aktif olduğu için gecikme süresini actual olarak alalım
                                'delay': int(delay_seconds),
                                'status': 'active'
                            })
                    except ValueError:
                        pass

        # TTO (Time To Own) kontrolü
        if sla_tto_id and sla_tto_id in slas:
            sla_tto = slas[sla_tto_id]
            sla_limit_seconds = convert_to_seconds(
                sla_tto['number_time'], 
                sla_tto['definition_time']
            )
            
            if time_to_own > 0 and time_to_own > sla_limit_seconds:
                has_violation = True
                delay = time_to_own - sla_limit_seconds
                violation_details.append({
                    'type': 'TTO',
                    'sla_name': sla_tto['name'],
                    'limit': sla_limit_seconds,
                    'actual': time_to_own,
                    'delay': delay,
                    'status': 'assigned'
                })
        
        # İstatistikleri güncelle
        if has_violation:
            stats[entity_name]['sla_violated'] += 1
        elif solve_delay_stat > 0 or time_to_own > 0:
            stats[entity_name]['sla_ok'] += 1
        else:
            # Aktif ticketlar için henüz ihlal yoksa buraya düşecek (Süresi devam edenler)
            stats[entity_name]['sla_active'] += 1
        
        # Ticket detaylarını kaydet
        ticket_info = {
            'id': ticket.get('id'),
            'name': ticket.get('name'),
            'date': ticket.get('date'),
            'solve_date': ticket.get('solvedate'),
            'sla_ttr_id': sla_ttr_id,
            'sla_tto_id': sla_tto_id,
            'solve_delay': solve_delay_stat,
            'time_to_own': time_to_own,
            'has_violation': has_violation,
            'violations': violation_details
        }
        
        stats[entity_name]['tickets'].append(ticket_info)
    
    return dict(stats)

def convert_to_seconds(number_time, definition_time):
    """SLA süresini saniyeye çevir"""
    if definition_time == 'minute':
        return number_time * 60
    elif definition_time == 'hour':
        return number_time * 3600
    elif definition_time == 'day':
        return number_time * 86400
    else:
        return number_time * 3600  # Default: hour

def print_report(stats):
    """Raporu konsola yazdır"""
    print("\n" + "="*120)
    print("SLA UYUM RAPORU - ENTITY BAZLI KIRILIM")
    print("="*140)
    print("{:<40} {:<10} {:<10} {:<12} {:<12} {:<12} {:<12}".format('Entity', 'Toplam', "SLA'lı", 'SLA Uygun', 'SLA İhlal', 'SLA Devam', 'İhlal Oranı'))
    print("-"*140)
    
    total_tickets = 0
    total_sla_tickets = 0
    total_ok = 0
    total_violated = 0
    total_active = 0
    
    # Entity'leri toplam ticket sayısına göre sırala
    sorted_stats = sorted(stats.items(), key=lambda x: x[1]['total'], reverse=True)
    
    for entity_name, data in sorted_stats:
        total = data['total']
        sla_total = data['sla_total']
        sla_ok = data['sla_ok']
        sla_violated = data['sla_violated']
        sla_active = data.get('sla_active', 0)
        violation_rate = (sla_violated / sla_total * 100) if sla_total > 0 else 0
        
        total_tickets += total
        total_sla_tickets += sla_total
        total_ok += sla_ok
        total_violated += sla_violated
        total_active += sla_active
        
        print(f"{entity_name:<40} {total:<10} {sla_total:<10} {sla_ok:<12} {sla_violated:<12} {sla_active:<12} {violation_rate:>10.2f}%")
    
    print("-"*140)
    total_violation_rate = (total_violated / total_sla_tickets * 100) if total_sla_tickets > 0 else 0
    print(f"{'TOPLAM':<40} {total_tickets:<10} {total_sla_tickets:<10} {total_ok:<12} {total_violated:<12} {total_active:<12} {total_violation_rate:>10.2f}%")
    print("="*140)

def export_to_csv(stats, filename):
    """CSV olarak export et"""
    import csv
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['Entity', 'Toplam Ticket', 'SLA\'lı Ticket', 'SLA Uygun', 'SLA İhlal', 'SLA Devam (Aktif)', 'İhlal Oranı %'])
        
        sorted_stats = sorted(stats.items(), key=lambda x: x[1]['total'], reverse=True)
        
        for entity_name, data in sorted_stats:
            total = data['total']
            sla_total = data['sla_total']
            sla_ok = data['sla_ok']
            sla_violated = data['sla_violated']
            sla_active = data.get('sla_active', 0)
            violation_rate = (sla_violated / sla_total * 100) if sla_total > 0 else 0
            
            writer.writerow([entity_name, total, sla_total, sla_ok, sla_violated, sla_active, f"{violation_rate:.2f}"])
    
    print(f"\n✓ Rapor CSV olarak kaydedildi: {filename}")

def export_detailed_csv(stats, slas, filename):
    """Detaylı CSV olarak export et (ticket bazında)"""
    import csv
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Entity', 
            'Ticket ID', 
            'Ticket Adı',
            'Açılış Tarihi',
            'Çözüm Tarihi',
            'SLA Durumu',
            'TTR İhlal',
            'TTR SLA Adı',
            'TTR Limit',
            'TTR Gerçek',
            'TTR Gecikme',
            'TTO İhlal',
            'TTO SLA Adı',
            'TTO Limit',
            'TTO Gerçek',
            'TTO Gecikme'
        ])
        
        # Entity'leri sırala
        sorted_stats = sorted(stats.items(), key=lambda x: x[1]['total'], reverse=True)
        
        for entity_name, data in sorted_stats:
            for ticket in data['tickets']:
                # TTR bilgileri
                ttr_violation = False
                ttr_sla_name = ''
                ttr_limit = ''
                ttr_actual = ''
                ttr_delay = ''
                
                # TTO bilgileri
                tto_violation = False
                tto_sla_name = ''
                tto_limit = ''
                tto_actual = ''
                tto_delay = ''
                
                # İhlal detaylarını kontrol et
                for violation in ticket.get('violations', []):
                    if violation['type'] == 'TTR':
                        ttr_violation = True
                        ttr_sla_name = violation['sla_name']
                        ttr_limit = format_duration(violation['limit'])
                        ttr_actual = format_duration(violation['actual'])
                        ttr_delay = format_duration(violation['delay'])
                    elif violation['type'] == 'TTO':
                        tto_violation = True
                        tto_sla_name = violation['sla_name']
                        tto_limit = format_duration(violation['limit'])
                        tto_actual = format_duration(violation['actual'])
                        tto_delay = format_duration(violation['delay'])
                
                # SLA bilgilerini al (ihlal yoksa)
                if not ttr_violation and ticket.get('sla_ttr_id') and ticket['sla_ttr_id'] in slas:
                    ttr_sla_name = slas[ticket['sla_ttr_id']]['name']
                    if ticket.get('solve_delay', 0) > 0:
                        ttr_actual = format_duration(ticket['solve_delay'])
                
                if not tto_violation and ticket.get('sla_tto_id') and ticket['sla_tto_id'] in slas:
                    tto_sla_name = slas[ticket['sla_tto_id']]['name']
                    if ticket.get('time_to_own', 0) > 0:
                        tto_actual = format_duration(ticket['time_to_own'])
                
                # Genel SLA durumu
                if ticket.get('has_violation'):
                    sla_status = 'İHLAL'
                elif ticket.get('solve_delay', 0) > 0 or ticket.get('time_to_own', 0) > 0:
                    sla_status = 'UYGUN'
                else:
                    sla_status = 'VERİ YOK'
                
                writer.writerow([
                    entity_name,
                    ticket.get('id', ''),
                    ticket.get('name', ''),
                    ticket.get('date', ''),
                    ticket.get('solve_date', ''),
                    sla_status,
                    'EVET' if ttr_violation else 'HAYIR',
                    ttr_sla_name,
                    ttr_limit,
                    ttr_actual,
                    ttr_delay,
                    'EVET' if tto_violation else 'HAYIR',
                    tto_sla_name,
                    tto_limit,
                    tto_actual,
                    tto_delay
                ])
    
    print(f"\n✓ Detaylı rapor CSV olarak kaydedildi: {filename}")

def format_duration(seconds):
    """Saniyeyi okunabilir formata çevir"""
    if seconds == 0:
        return ''
    
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    
    parts = []
    if days > 0:
        parts.append(f"{days}g")
    if hours > 0:
        parts.append(f"{hours}s")
    if minutes > 0:
        parts.append(f"{minutes}d")
    
    return ' '.join(parts) if parts else f"{seconds}sn"

def export_to_excel(stats, filename):
    """Excel olarak export et"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SLA Uyum Raporu"
        
        # Başlıklar
        headers = ['Entity', 'Toplam Ticket', 'SLA Uygun', 'SLA İhlal', 'İhlal Oranı %']
        ws.append(headers)
        
        # Başlık stilini ayarla
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Verileri ekle
        sorted_stats = sorted(stats.items(), key=lambda x: x[1]['total'], reverse=True)
        
        for entity_name, data in sorted_stats:
            total = data['total']
            sla_ok = data['sla_ok']
            sla_violated = data['sla_violated']
            violation_rate = (sla_violated / total * 100) if total > 0 else 0
            
            ws.append([entity_name, total, sla_ok, sla_violated, violation_rate])
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        wb.save(filename)
        print(f"\n✓ Rapor Excel olarak kaydedildi: {filename}")
        
    except ImportError:
        print("\n⚠ Excel export için 'openpyxl' kütüphanesi gerekli:")
        print("  pip install openpyxl")

def main():
    parser = argparse.ArgumentParser(description='GLPI SLA Uyum Raporu')
    parser.add_argument('--start-date', required=True, help='Başlangıç tarihi (YYYY-MM-DD)')
    parser.add_argument('--end-date', required=True, help='Bitiş tarihi (YYYY-MM-DD)')
    parser.add_argument('--export', choices=['csv', 'excel'], help='Export formatı')
    
    args = parser.parse_args()
    
    # Config yükle
    config = load_config()
    
    # Session başlat
    session_token = init_session(config)
    
    try:
        # Entity'leri çek
        entities = fetch_all_entities(config, session_token)
        
        # SLA tanımlarını çek
        slas = fetch_all_slas(config, session_token)
        
        # SLA'lı ticketları çek
        tickets = fetch_tickets(config, session_token, args.start_date, args.end_date)
        
        if not tickets:
            print("\n⚠ Belirtilen tarih aralığında SLA'lı ticket bulunamadı!")
            return
        
        # Analiz et
        stats = analyze_sla_compliance(tickets, entities, slas)
        
        # Raporu göster
        print_report(stats)
        
        # Export
        if args.export == 'csv':
            # Özet rapor
            summary_filename = f"sla_report_summary_{args.start_date}_{args.end_date}.csv"
            export_to_csv(stats, summary_filename)
            
            # Detaylı rapor
            detailed_filename = f"sla_report_detailed_{args.start_date}_{args.end_date}.csv"
            export_detailed_csv(stats, slas, detailed_filename)
            
        elif args.export == 'excel':
            filename = f"sla_report_{args.start_date}_{args.end_date}.xlsx"
            export_to_excel(stats, filename)
        
    finally:
        # Session kapat
        kill_session(config, session_token)

if __name__ == '__main__':
    main()
